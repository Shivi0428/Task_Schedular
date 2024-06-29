import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime

# Sample data
sample_data = [
    ["Design project layout", "Not Started", "", ""],
    ["Write project documentation", "Under Process", "2024-06-27 14:45:00", "Documentation is halfway done."],
    ["Implement user authentication", "Completed", "2024-06-25 11:30:00", ""],
    ["Setup database schema", "Completed", "2024-06-26 09:20:00", ""],
    ["Conduct code review", "Not Started", "", ""],
    ["Test application features", "Under Process", "2024-06-28 10:00:00", "Initial testing started, some bugs found."],
    ["Deploy application to production", "Not Started", "", ""],
    ["Optimize application performance", "Completed", "2024-06-24 08:45:00", ""],
    ["Gather user feedback", "Not Started", "", ""],
    ["Schedule team meeting", "Completed", "2024-06-23 15:00:00", ""]]

try:
    wb = openpyxl.load_workbook("TaskScheduler.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task Scheduler"
    headers = ["Task", "Status", "Timestamp", "Notes"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    for row_data in sample_data:
        ws.append(row_data)
    wb.save("TaskScheduler.xlsx")

def add_task(task_description):
    if not task_description.strip():
        messagebox.showerror("Error", "Task description cannot be empty")
        return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=task_description)
    ws.cell(row=row, column=2, value="Not Started")
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value="")
    wb.save("TaskScheduler.xlsx")
    list_tasks()
    task_entry.delete(0, tk.END)

def update_task_status(status):
    selected_task = tree.focus()
    if not selected_task:
        messagebox.showerror("Error", "No task selected")
        return
    task_row = int(selected_task) + 1
    ws.cell(row=task_row, column=2, value=status)
    ws.cell(row=task_row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    wb.save("TaskScheduler.xlsx")
    list_tasks()

def list_tasks():
    tree.delete(*tree.get_children())  # Clear existing items
    for row_id, row in enumerate(ws.iter_rows(min_row=2, max_row=min(ws.max_row, 31), min_col=1, max_col=4, values_only=True), start=2):
        task_description, status, timestamp, notes = row
        tree.insert("", tk.END, iid=row_id, values=(task_description, status, timestamp, notes))
        if status == "Not Started":
            tree.item(row_id, tags=("not_started",))
        elif status == "Under Process":
            tree.item(row_id, tags=("under_process",))
        elif status == "Completed":
            tree.item(row_id, tags=("completed",))
    tree.yview_moveto(1.0)

def delete_task():
    selected_task = tree.focus()
    if not selected_task:
        messagebox.showerror("Error", "No task selected")
        return
    task_row = int(selected_task) + 1
    ws.delete_rows(task_row)
    wb.save("TaskScheduler.xlsx")
    list_tasks()

def add_notes():
    selected_task = tree.focus()
    if not selected_task:
        messagebox.showerror("Error", "No task selected")
        return
    task_row = int(selected_task) + 1
    notes_value = notes_entry.get()
    ws.cell(row=task_row, column=4, value=notes_value)
    wb.save("TaskScheduler.xlsx")
    list_tasks()
    notes_entry.delete(0, tk.END)

# GUI setup
root = tk.Tk()
root.title("Task Scheduler")
root.configure(bg="#20bebe")
root.geometry("900x600")  # Adjusted window size

# Task entry
tk.Label(root, text="Enter Task:", bg="#20bebe").grid(row=0, column=0, padx=10, pady=10)
task_entry = tk.Entry(root, width=80)
task_entry.grid(row=0, column=1, padx=10, pady=10, columnspan=3)
tk.Button(root, text="Add Task", command=lambda: add_task(task_entry.get())).grid(row=0, column=4, padx=10, pady=10)

# Treeview for tasks with scrollbars
columns = ("Task", "Status", "Timestamp", "Notes")
tree = ttk.Treeview(root, columns=columns, show="headings")
tree.grid(row=1, column=0, columnspan=5, padx=10, pady=10, sticky="nsew")

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=200, anchor="center")

tree.tag_configure("not_started", background="orange")
tree.tag_configure("under_process", background="lightblue")
tree.tag_configure("completed", background="lightgreen")

# Scrollbars
vsb = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
vsb.grid(row=1, column=5, rowspan=2, sticky="ns")
hsb = ttk.Scrollbar(root, orient="horizontal", command=tree.xview)
hsb.grid(row=3, column=0, columnspan=5, sticky="ew")
tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

# Task status buttons


# Notes entry and add notes button
tk.Label(root, text="Notes:", bg="#20bebe").grid(row=10, column=0, padx=10, pady=10)
notes_entry = tk.Entry(root, width=80)
notes_entry.grid(row=10, column=1, columnspan=3, padx=10, pady=10)
tk.Button(root, text="Add Notes", command=add_notes).grid(row=10, column=4, padx=10, pady=10)
tk.Button(root, text="Mark as Not Started", command=lambda: update_task_status("Not Started")).grid(row=12, column=0, padx=10, pady=10)
tk.Button(root, text="Mark as Under Process", command=lambda: update_task_status("Under Process")).grid(row=12, column=1, padx=10, pady=10)
tk.Button(root, text="Mark as Completed", command=lambda: update_task_status("Completed")).grid(row=12, column=2, padx=10, pady=10)

# Delete task button
tk.Button(root, text="Delete Task", command=delete_task).grid(row=12, column=4, padx=10, pady=10)
# Marquee with motivational quote or link
quote = "“None can destroy iron, but its own rust can“"
marquee_label = tk.Label(root, text=quote, bg="#20bebe", font=("Helvetica", 12, "italic"))
marquee_label.grid(row=14, column=0, columnspan=5, pady=10)
marquee_label.config(fg="white")  # Adjust text color if needed

# Initial task list load
list_tasks()

# Start the GUI loop
root.mainloop()
